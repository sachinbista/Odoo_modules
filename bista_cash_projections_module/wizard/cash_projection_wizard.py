from time import strftime
from odoo import models, fields, _
import base64
from io import BytesIO
from datetime import timedelta, datetime
from openpyxl import load_workbook
from openpyxl.styles import Font, Border, Side, GradientFill, Alignment
import os
from odoo.modules.module import get_module_resource
from odoo.tools import date_utils


class CashProjectionWizard(models.TransientModel):
    _name = 'cash.projection.wizard'
    _description = 'Cash Projection Wizard'

    start_date = fields.Date(string='Start Date', required=True)
    end_date = fields.Date(string='End Date', required=True)

    def _check_dates(self):
        for record in self:
            if record.start_date > record.end_date:
                raise models.ValidationError(_('Start Date cannot be greater than End Date.'))

    def delete_column_by_name(self, sheet, column_name):
        column_index = None
        for col in range(1, sheet.max_column + 1):
            if sheet.cell(row=1, column=col).value == column_name:
                column_index = col
                break

        if column_index is not None:
            for row in sheet.iter_rows():
                del row[column_index - 1]

    def get_all_year_week_records(self, start_date):
        all_year_week = {}
        current_date = start_date
        start_of_week = start_date - timedelta(days=start_date.weekday())

        for data in range(52):
            year, current_week, _ = current_date.isocalendar()

            if current_week not in all_year_week:
                week_start_date = start_of_week
                week_end_date = start_of_week + timedelta(days=6)
                all_year_week[current_week] = {'start_date': week_start_date, 'end_date': week_end_date}

            start_of_week += timedelta(days=7)
            current_date = start_of_week

        return all_year_week

    def calculate_cash_projection_report(self):
        self._check_dates()

        ir_attachment_obj = self.env['ir.attachment']

        attach_id = ir_attachment_obj.search([('name', '=', 'Original Cash Projections.xlsx')], order='id asc', limit=1)

        if not attach_id:
            text_file_path = get_module_resource('bista_cash_projections_module', 'data/', '')
            file_path = os.path.join(text_file_path, 'Original_Cash_Projections.xlsx')

            with open(file_path, 'rb') as file:
                file_data = file.read()

            attachment_vals = {
                'datas': base64.b64encode(file_data),
                'name': 'Original Cash Projections.xlsx',
            }

            attach_id = ir_attachment_obj.create(attachment_vals)
        attach_id.ensure_one()

        file_stream = BytesIO(base64.b64decode(attach_id.datas))
        workbook = load_workbook(file_stream)
        sheet_name = 'Sheet1'
        sheet = workbook[sheet_name]

        self.delete_column_by_name(sheet, "ColumnToDelete")

        start_date = self.start_date
        end_date = self.end_date
        date_str = 'Start Date :- ' + start_date.strftime('%d/%m/%y') + '\n' + 'End Date :- ' + end_date.strftime(
            '%d/%m/%y')
        sheet.merge_cells('B2:B3')
        sheet.cell(row=2, column=2, value=date_str)

        account_move_flow = {}
        sale_order_flow = {}
        account_move_payable_flow = {}
        purchase_order_flow = {}
        cash_receipt_total = {}

        week_start_dates = {}
        week_end_dates = {}

        if start_date and end_date:
            current_date = start_date
            col_num = 4  # Start from column F (6th column)
            row_num = 6  # Start from row 6

            start_date = current_date - timedelta(days=current_date.weekday())
            end_date = start_date + timedelta(days=6)
            count = 1

            all_year_week = self.get_all_year_week_records(self.start_date)

            col_num = 5

            for key, value in all_year_week.items():
                week_start = value['start_date']
                week_end = value['end_date']
                cell_value = f'{week_start.strftime("%d/%m/%y")} - {week_end.strftime("%d/%m/%y")}'
                font_style = Font(bold=True, size=11)
                align_style = Alignment(horizontal="center")
                row_num = 2

                sheet.cell(row=2, column=col_num, value=cell_value).font = font_style
                sheet.cell(row=2, column=col_num).alignment = align_style

                col_num += 1

            while end_date <= self.end_date:
                year, current_week, _ = current_date.isocalendar()

                if count == 1:
                    start_date = self.start_date

                if end_date > self.end_date:
                    end_date = self.end_date

                if current_week not in account_move_flow:
                    account_move_flow[current_week] = 0
                if current_week not in sale_order_flow:
                    sale_order_flow[current_week] = 0
                if current_week not in account_move_payable_flow:
                    account_move_payable_flow[current_week] = 0
                if current_week not in purchase_order_flow:
                    purchase_order_flow[current_week] = 0
                if current_week not in cash_receipt_total:
                    cash_receipt_total[current_week] = 0

                col_num += 1

                # Query for account_move
                query_account_move = """
                        SELECT 
                        SUM(am.amount_residual) AS total_amount,
                        DATE_PART('week', am.invoice_date_due) AS week
                        FROM 
                        account_move am
                        WHERE 
                        am.invoice_date_due BETWEEN %s AND %s
                        AND am.move_type = 'out_invoice'
                        AND am.state = 'posted'
                        GROUP BY 
                        DATE_PART('week', am.invoice_date_due)
                        ORDER BY 
                        week;
                    """

                self.env.cr.execute(query_account_move, (start_date, end_date))
                data_rows_out_invoice = self.env.cr.fetchall()

                for data_row in data_rows_out_invoice:
                    account_move_flow[current_week] += data_row[0]
                    cash_receipt_total[current_week] += data_row[0]

                # Query for sale order
                query_sale_order = """
                    SELECT
                        SUM((sol.product_uom_qty - sol.qty_delivered) * sol.price_unit) AS total_amount,
                        CASE
                            WHEN so.payment_term_id IS NULL THEN DATE_PART('week', sp.scheduled_date) 
                            WHEN so.payment_term_id IS NOT NULL THEN DATE_PART('week', sp.scheduled_date + interval '1 day' * (line.days - 1))
                            ELSE NULL
                        END AS week_number,
                        so.name,
                        sp.scheduled_date AS picking_scheduled_date,
                        line.days AS payment_term_days,
                        sp.scheduled_date + interval '1 day' * (line.days - 1) AS due_date,
                        sol.product_id,
                        sol.product_uom_qty,
                        sp.state,
                        op.code
                    FROM
                        sale_order AS so
                    JOIN
                        stock_picking AS sp ON so.id = sp.sale_id
                    LEFT JOIN
                        account_payment_term_line AS line ON line.payment_id = so.payment_term_id
                    JOIN
                        sale_order_line AS sol ON so.id = sol.order_id
                    LEFT JOIN
                        stock_picking_type as op on op.id = sp.picking_type_id
                    WHERE
                        CASE
                            WHEN so.payment_term_id IS NULL THEN sp.scheduled_date::date BETWEEN %s AND %s  
                            WHEN so.payment_term_id IS NOT NULL THEN (sp.scheduled_date + interval '1 day' * (line.days - 1))::date BETWEEN %s AND %s   
                        END
                        AND so.state = 'sale'
                        AND sp.state IN ('waiting', 'confirmed', 'assigned')
                        AND op.code = 'outgoing'
                    GROUP BY
                        week_number, so.name, line.value, sp.scheduled_date, line.days, sol.product_id, sol.product_uom_qty, sp.state, op.code
                    ORDER BY
                        week_number;
                         """
                self.env.cr.execute(query_sale_order, (start_date, end_date, start_date, end_date))
                sale_order_data = self.env.cr.fetchall()

                for data_row in sale_order_data:
                    if data_row[0] is not None:
                        sale_order_flow[current_week] += data_row[0]
                        cash_receipt_total[current_week] += data_row[0]
                        # if data_row[1] is not None and data_row[0] is not None:
                        #     if int(data_row[1]) in sale_order_flow:
                        #         account_move_flow[int(data_row[1])] += float(data_row[0])

                total_amount = sum(
                    sale_order[0] if sale_order[0] is not None else 0 for sale_order in sale_order_data)

                sum_data_out_invoice = sum(data_row[0] for data_row in data_rows_out_invoice if data_row[0] is not None)
                sum_data_sale_order = total_amount if total_amount is not None else 0

                total_sum = sum_data_out_invoice + sum_data_sale_order
                # sheet.cell(row=15, column=5, value=total_sum)

                # Query for account_move_payable
                query_account_move_payable = """
                    SELECT						     
                        SUM(am.amount_residual) AS total_amount,
                        DATE_PART('week', am.invoice_date_due) AS week
                    FROM 
                        account_move am
                    WHERE 
                        am.invoice_date_due BETWEEN %s AND %s
                        AND am.move_type = 'in_invoice'
                        AND am.state = 'posted'
                    GROUP BY 
                        DATE_PART('week', am.invoice_date_due)
                    ORDER BY 
                        week;
                    """
                self.env.cr.execute(query_account_move_payable, (start_date, end_date))
                account_move_payble_total = self.env.cr.fetchall()

                for data_row in account_move_payble_total:
                    if int(data_row[1]) in account_move_payable_flow:
                        account_move_payable_flow[int(data_row[1])] += data_row[0]

                # Query for purchase order
                query_purchase_order = """
                    SELECT
                        SUM(po.amount_total) AS total_amount,
                    CASE
                        WHEN po.payment_term_id IS NULL THEN DATE_PART('week', sp.scheduled_date) 
                        WHEN po.payment_term_id IS NOT NULL THEN DATE_PART('week', sp.scheduled_date + interval '1 day' * (line.days - 1))
                        ELSE NULL
                        END AS week_number,
                                po.name,
                                sp.scheduled_date AS receiving_scheduled_date,
                                line.days AS payment_term_days,
                                sp.scheduled_date + interval '1 day' * (line.days - 1) AS due_date
                            FROM
                                purchase_order AS po
                            JOIN
                               stock_picking AS sp ON sp.group_id = po.group_id
                            LEFT JOIN
                                account_payment_term_line AS line ON line.payment_id = po.payment_term_id
                            LEFT JOIN
	                            stock_picking_type as op on op.id = sp.picking_type_id
                            WHERE
                                CASE
                                    WHEN po.payment_term_id IS NULL THEN sp.scheduled_date::date BETWEEN %s AND %s 
                                    WHEN po.payment_term_id IS NOT NULL THEN (sp.scheduled_date + interval '1 day' * (line.days - 1))::date BETWEEN %s AND %s 
                            END
                                AND po.state = 'purchase'
                                AND sp.state IN ('waiting', 'confirmed', 'assigned')
                                AND op.code = 'incoming'
                            GROUP BY
                                week_number, po.name, line.value, sp.scheduled_date, line.days ,sp.state, op.code
                            ORDER BY
                                week_number
                    """
                self.env.cr.execute(query_purchase_order, (start_date, end_date, start_date, end_date))
                purchase_order_total_records = self.env.cr.fetchall()

                for data_row in purchase_order_total_records:
                    if int(data_row[1]) in purchase_order_flow:
                        purchase_order_flow[int(data_row[1])] += data_row[0]

                col_num += 1

                week_start_dates[current_week] = start_date
                week_end_dates[current_week] = end_date

                # current_date += timedelta(days=7)
                current_date = date_utils.start_of(current_date, "week") + timedelta(days=7)

                start_date = end_date + timedelta(days=1)
                end_date = start_date + timedelta(days=6)

                count += 1

            col_num = 5
            for key, value in account_move_flow.items():
                value = round(account_move_flow.get(key, 0), 2)
                align_style = Alignment(horizontal="center")
                sheet.cell(row=8, column=col_num, value=value).alignment = align_style

                col_num += 1

            col_num = 5
            for key, value in sale_order_flow.items():
                value = round(sale_order_flow.get(key, 0), 2)
                align_style = Alignment(horizontal="center")
                sheet.cell(row=9, column=col_num, value=value).alignment = align_style

                col_num += 1

            col_num = 5
            for key, value in account_move_payable_flow.items():
                value = round(account_move_payable_flow.get(key, 0), 2)
                align_style = Alignment(horizontal="center")
                sheet.cell(row=18, column=col_num, value=value).alignment = align_style

                col_num += 1

            col_num = 5
            for key, value in cash_receipt_total.items():
                font_style = Font(bold=True, size=11, name='cambria')
                value = round(cash_receipt_total.get(key, 0), 2)
                align_style = Alignment(horizontal="center")
                sheet.cell(row=15, column=col_num, value=value).font = font_style
                sheet.cell(row=15, column=col_num, value=value).alignment = align_style

                col_num += 1

            col_num = 5
            for key, value in purchase_order_flow.items():
                value = round(purchase_order_flow.get(key, 0), 2)
                align_style = Alignment(horizontal="center")
                sheet.cell(row=19, column=col_num, value=value).alignment = align_style

                col_num += 1

            row_num += 1

        modified_file_stream = BytesIO()
        workbook.save(modified_file_stream)
        attach_id.write({'datas': base64.b64encode(modified_file_stream.getvalue())})

        return {
            'type': 'ir.actions.act_url',
            'url': '/web/content/%s/%s' % (attach_id.id, attach_id.name),
            'target': 'new',
            'nodestroy': False,
        }
